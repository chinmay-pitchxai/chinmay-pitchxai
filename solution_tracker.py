#!/usr/bin/env python3
"""Solution versioning bug tracker backed by an Excel workbook."""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

BASE_DIR = Path(__file__).resolve().parent
TRACKER_PATH = BASE_DIR / "SOLUTION_TRACKER.xlsx"
REPORT_PATH = BASE_DIR / "DEBUGGING_REPORT.xlsx"
REPORT_SHEET = "Solution Versions"

BUGS_HEADER = ["BugID", "Problem", "Category", "Tags", "Created (UTC ISO)"]
SOLUTIONS_HEADER = [
    "BugID",
    "Problem",
    "Category",
    "Tags",
    "Version",
    "Solution",
    "Verdict",
    "Notes",
    "Chosen",
    "Timestamp (UTC ISO)",
]
MAX_VERSIONS = 15
VERDICTS = ("", "works", "partial", "failed")


def now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def load_tracker() -> Workbook:
    if TRACKER_PATH.is_file():
        return load_workbook(TRACKER_PATH)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def ensure_sheets(wb: Workbook) -> None:
    if "Bugs" not in wb.sheetnames:
        wb.create_sheet("Bugs").append(BUGS_HEADER)
    if "Solutions" not in wb.sheetnames:
        wb.create_sheet("Solutions").append(SOLUTIONS_HEADER)


def style_headers(wb: Workbook) -> None:
    bold = Font(bold=True)
    for sheet in ("Bugs", "Solutions"):
        for cell in wb[sheet][1]:
            cell.font = bold


def next_bug_id(wb: Workbook) -> str:
    max_num = 0
    for row in wb["Bugs"].iter_rows(min_row=2, values_only=True):
        m = re.fullmatch(r"BUG-(\d+)", str(row[0] or "").strip().upper())
        if m:
            max_num = max(max_num, int(m.group(1)))
    return f"BUG-{max_num + 1:03d}"


def find_bug(wb: Workbook, bug_id: str) -> tuple[str, str, str] | None:
    key = bug_id.strip().upper()
    for row in wb["Bugs"].iter_rows(min_row=2, values_only=True):
        if str(row[0] or "").strip().upper() == key:
            return (str(row[1] or ""), str(row[2] or ""), str(row[3] or ""))
    return None


def find_bug_by_problem(wb: Workbook, problem: str) -> str | None:
    key = problem.strip().lower()
    for row in wb["Bugs"].iter_rows(min_row=2, values_only=True):
        if str(row[1] or "").strip().lower() == key:
            return str(row[0])
    return None


def count_versions(wb: Workbook, bug_id: str) -> int:
    key = bug_id.strip().upper()
    n = 0
    for row in wb["Solutions"].iter_rows(min_row=2, values_only=True):
        if str(row[0] or "").strip().upper() == key and str(row[4] or "").strip():
            n += 1
    return n


def version_row_index(wb: Workbook, bug_id: str, version: str) -> int | None:
    key = bug_id.strip().upper()
    for row in wb["Solutions"].iter_rows(min_row=2):
        if str(row[0].value or "").strip().upper() == key and str(row[4].value or "").strip().upper() == version:
            return row[0].row
    return None


def normalize_version(raw: str) -> str | None:
    v = raw.strip().upper()
    if re.fullmatch(r"S\d+", v):
        return v
    if re.fullmatch(r"\d+", v):
        return "S" + v
    return None


def cmd_new(args: argparse.Namespace) -> int:
    wb = load_tracker()
    ensure_sheets(wb)
    existing = find_bug_by_problem(wb, args.problem)
    if existing:
        print(f"Bug already exists: {existing}")
        return 0
    bug_id = next_bug_id(wb)
    wb["Bugs"].append([bug_id, args.problem.strip(), args.category.strip(), args.tag.strip(), now_iso()])
    style_headers(wb)
    wb.save(TRACKER_PATH)
    print(f"Created {bug_id}")
    return 0


def cmd_add(args: argparse.Namespace) -> int:
    wb = load_tracker()
    ensure_sheets(wb)
    bug_id = args.bug.strip().upper()
    bug = find_bug(wb, bug_id)
    if bug is None:
        print(f"Unknown bug ID: {args.bug}", file=sys.stderr)
        return 1
    verdict = args.verdict.strip().lower()
    if verdict not in VERDICTS:
        print(f"Invalid verdict: {args.verdict} (use works|partial|failed)", file=sys.stderr)
        return 1
    n = count_versions(wb, bug_id)
    if n >= MAX_VERSIONS:
        print(f"Bug {bug_id} already has {MAX_VERSIONS} solution versions (max reached)", file=sys.stderr)
        return 1
    problem, category, tags = bug
    version = f"S{n + 1}"
    wb["Solutions"].append(
        [bug_id, problem, category, tags, version, args.solution.strip(), verdict, args.notes.strip(), "", now_iso()]
    )
    style_headers(wb)
    wb.save(TRACKER_PATH)
    print(f"Added {version} for {bug_id} (total {n + 1}/{MAX_VERSIONS})")
    return 0


def cmd_choose(args: argparse.Namespace) -> int:
    wb = load_tracker()
    ensure_sheets(wb)
    bug_id = args.bug.strip().upper()
    if find_bug(wb, bug_id) is None:
        print(f"Unknown bug ID: {args.bug}", file=sys.stderr)
        return 1
    version = normalize_version(args.version)
    if version is None:
        print(f"Invalid version: {args.version} (expected e.g. S3)", file=sys.stderr)
        return 1
    if version_row_index(wb, bug_id, version) is None:
        print(f"No version {version} for {bug_id}", file=sys.stderr)
        return 1
    ws = wb["Solutions"]
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value or "").strip().upper() == bug_id:
            row[8].value = "TRUE" if row[0].row == version_row_index(wb, bug_id, version) else ""
    wb.save(TRACKER_PATH)
    print(f"Marked {version} of {bug_id} as CHOSEN")
    return 0


def cmd_status(args: argparse.Namespace) -> int:
    wb = load_tracker()
    ensure_sheets(wb)
    bugs = []
    for row in wb["Bugs"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            bugs.append((str(row[0]).strip(), str(row[1] or "").strip()))
    if not bugs:
        print("No bugs tracked yet.")
        return 0
    counts: dict[str, int] = {}
    chosen: dict[str, str] = {}
    for row in wb["Solutions"].iter_rows(min_row=2, values_only=True):
        if not str(row[4] or "").strip():
            continue
        bid = str(row[0] or "").strip().upper()
        counts[bid] = counts.get(bid, 0) + 1
        if str(row[8] or "").strip().upper() == "TRUE":
            chosen[bid] = str(row[4])
    header = ["BugID", "Problem", "Solutions", "Chosen"]
    lines = [[bid, problem, str(counts.get(bid.upper(), 0)), chosen.get(bid.upper(), "")] for bid, problem in bugs]
    widths = [max(len(str(r[i])) for r in [header] + lines) for i in range(4)]
    widths[1] = min(widths[1], 40)
    fmt = "  ".join("{:<" + str(w) + "}" for w in widths)
    print(fmt.format(*header))
    print("-" * (sum(widths) + 2 * (len(widths) - 1)))
    for line in lines:
        print(fmt.format(line[0], line[1][:widths[1]], line[2], line[3]))
    return 0


def cmd_export(args: argparse.Namespace) -> int:
    if not TRACKER_PATH.is_file():
        print(f"SOLUTION_TRACKER.xlsx not found - run 'new' or 'add' first", file=sys.stderr)
        return 1
    if not REPORT_PATH.is_file():
        print(f"DEBUGGING_REPORT.xlsx not found at {REPORT_PATH}", file=sys.stderr)
        return 1
    src = load_workbook(TRACKER_PATH)
    ensure_sheets(src)
    dst = load_workbook(REPORT_PATH)
    if REPORT_SHEET in dst.sheetnames:
        del dst[REPORT_SHEET]
    out = dst.create_sheet(REPORT_SHEET)
    rows = list(src["Solutions"].iter_rows(values_only=True))
    for row in rows:
        out.append(row)
    for cell in out[1]:
        cell.font = Font(bold=True)
    dst.save(REPORT_PATH)
    print(f"Exported {max(len(rows) - 1, 0)} solution rows -> {REPORT_PATH} [{REPORT_SHEET}]")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="command", required=True)

    p_new = sub.add_parser("new", help="Create a new bug entry and print its BugID")
    p_new.add_argument("--problem", required=True, help="Problem description")
    p_new.add_argument("--category", default="", help="Optional category")
    p_new.add_argument("--tag", default="", help="Optional comma-separated tags")
    p_new.set_defaults(func=cmd_new)

    p_add = sub.add_parser("add", help="Append the next solution version (S1, S2, ...)")
    p_add.add_argument("--bug", required=True, help="BugID, e.g. BUG-001")
    p_add.add_argument("--solution", required=True, help="Solution text for this attempt")
    p_add.add_argument("--verdict", default="", help="works | partial | failed")
    p_add.add_argument("--notes", default="", help="Optional notes")
    p_add.set_defaults(func=cmd_add)

    p_choose = sub.add_parser("choose", help="Mark a version as CHOSEN (canonical)")
    p_choose.add_argument("--bug", required=True, help="BugID, e.g. BUG-001")
    p_choose.add_argument("--version", required=True, help="Version, e.g. S3 or 3")
    p_choose.set_defaults(func=cmd_choose)

    sub.add_parser("status", help="Summary of all bugs").set_defaults(func=cmd_status)
    sub.add_parser("export", help="Copy tracker into DEBUGGING_REPORT.xlsx").set_defaults(func=cmd_export)

    args = parser.parse_args()
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())