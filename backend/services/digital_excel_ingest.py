"""Idempotent Sandbox 1.2 ingestion for a continuously updated Excel/CSV file."""

from __future__ import annotations

import asyncio
import csv
import hashlib
from datetime import datetime, timezone
from pathlib import Path

from loguru import logger

from config import settings
from core.phone_norm import norm_phone_str
from core.storage import _bulk_add_leads_sync, _get_conn

PHONE_HEADERS = ("phone", "mobile", "contact", "whatsapp", "number", "telephone", "tel")
NAME_HEADERS = ("name", "full name", "customer", "lead", "contact name")
EMAIL_HEADERS = ("email", "mail")
COMPANY_HEADERS = ("company", "project", "organization", "business")


def _pick(headers: list[str], candidates: tuple[str, ...]) -> str | None:
    normalized = {h.strip().lower().replace("_", " "): h for h in headers}
    for candidate in candidates:
        if candidate in normalized:
            return normalized[candidate]
    for normalized_header, original in normalized.items():
        if any(candidate in normalized_header for candidate in candidates):
            return original
    return None


def read_digital_rows(path: Path, sheet_name: str = "") -> list[dict]:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        import openpyxl

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
        raw = list(worksheet.iter_rows(values_only=True))
        workbook.close()
        if not raw:
            return []
        headers = [str(value or f"column_{index + 1}").strip() for index, value in enumerate(raw[0])]
        records = [dict(zip(headers, row)) for row in raw[1:]]
    elif suffix in (".csv", ".tsv"):
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            records = list(csv.DictReader(handle, delimiter=delimiter))
        headers = list(records[0]) if records else []
    else:
        raise ValueError("Digital feed must be .xlsx, .xlsm, .csv, or .tsv")

    phone_col = _pick(headers, PHONE_HEADERS)
    if not phone_col:
        raise ValueError("Digital feed needs a phone/mobile/contact-number column")
    name_col = _pick(headers, NAME_HEADERS)
    email_col = _pick(headers, EMAIL_HEADERS)
    company_col = _pick(headers, COMPANY_HEADERS)
    mapped = {column for column in (phone_col, name_col, email_col, company_col) if column}
    leads: list[dict] = []
    seen: set[str] = set()
    for row in records:
        phone = norm_phone_str(str(row.get(phone_col) or ""))
        if not phone or phone in seen:
            continue
        seen.add(phone)
        extra = {str(k): str(v) for k, v in row.items() if k not in mapped and v not in (None, "")}
        leads.append({
            "name": str(row.get(name_col) or "Unknown") if name_col else "Unknown",
            "phone": phone,
            "email": str(row.get(email_col) or "") if email_col else "",
            "company": str(row.get(company_col) or "") if company_col else "",
            "source": "digital",
            "sandbox": 1,
            "upload_source": path.name,
            "extra": extra,
        })
    return leads


def ingest_digital_file(path: Path, *, role: str, sheet_name: str = "") -> dict:
    from core.state import normalize_console_role

    role = normalize_console_role(role)
    leads = read_digital_rows(path, sheet_name)
    saved, duplicates, dnc_blocked = _bulk_add_leads_sync(role, leads)
    conn = _get_conn()
    queued = 0
    from core.orchestration_service import schedule_job
    from core.workflow_models import JobType

    feed_id = hashlib.sha256(str(path.resolve()).encode("utf-8")).hexdigest()[:16]
    now = datetime.now(timezone.utc)
    for lead in leads:
        row = conn.execute(
            "SELECT id FROM leads WHERE role=? AND phone=? ORDER BY id DESC LIMIT 1",
            (role, lead["phone"]),
        ).fetchone()
        if not row:
            continue
        try:
            schedule_job(
                conn,
                lead_id=int(row[0]),
                job_type=JobType.FRESH_CALL,
                source="digital",
                due_at=now,
                key=f"digital-excel:{feed_id}:{int(row[0])}",
                attempt=1,
                source_type="digital_excel",
                source_id=feed_id,
                payload={"filename": path.name, "sub_sandbox": "1.2"},
            )
            queued += 1
        except Exception as exc:
            # Existing idempotency keys are expected on every later poll.
            if "UNIQUE" not in str(exc).upper():
                logger.warning("Digital Excel queue skipped lead {}: {}", row[0], exc)
    return {"rows": len(leads), "saved": saved, "duplicates": duplicates, "dnc_blocked": dnc_blocked, "queued": queued}


def ingest_digital_rows(rows: list[dict], *, broker_id: str) -> dict:
    """Normalize Google Sheet rows and reuse the same idempotent P3 queue."""
    from core.orchestration_service import schedule_job
    from core.workflow_models import JobType
    leads = []
    for item in rows:
        phone = norm_phone_str(str(item.get("phone") or ""))
        if not phone:
            continue
        leads.append({"name": str(item.get("name") or "Unknown"), "phone": phone,
                      "email": str(item.get("email") or ""), "details": str(item.get("notes") or ""),
                      "source": "digital", "sandbox": 1,
                      "extra": {"broker_id": broker_id, "sub_sandbox": "1.2"}})
    role = "sales_1"
    saved, duplicates, dnc_blocked = _bulk_add_leads_sync(role, leads)
    conn = _get_conn(); queued = 0
    for lead in leads:
        row = conn.execute("SELECT id FROM leads WHERE role=? AND phone=? ORDER BY id DESC LIMIT 1", (role, lead["phone"])).fetchone()
        if not row: continue
        try:
            schedule_job(conn, lead_id=int(row[0]), job_type=JobType.FRESH_CALL,
                         source="digital", due_at=datetime.now(timezone.utc),
                         key=f"google-sheet:{broker_id}:{int(row[0])}", attempt=1,
                         source_type="google_sheets", source_id=broker_id,
                         payload={"broker_id": broker_id, "sub_sandbox": "1.2"})
            queued += 1
        except Exception as exc:
            if "UNIQUE" not in str(exc).upper(): logger.warning("Sheet queue failed: {}", exc)
    return {"rows": len(rows), "saved": saved, "duplicates": duplicates, "dnc_blocked": dnc_blocked, "queued": queued}


async def digital_excel_watcher() -> None:
    configured = Path(settings.digital_excel_path).expanduser()
    last_signatures: dict[str, tuple[int, int]] = {}
    logger.info("Sandbox 1.2 digital Excel watcher configured for {}", configured)
    while True:
        try:
            # Self-heal: keep the feed directory present so channel partners can
            # drop files at any time without operator intervention.
            if not configured.exists():
                try:
                    configured.mkdir(parents=True, exist_ok=True)
                    logger.info("Auto-created digital Excel feed path: {}", configured)
                except Exception as _mkdir_exc:
                    logger.warning("Could not create digital Excel feed path {}: {}", configured, _mkdir_exc)
            
            files = (
                sorted(
                    (p for p in configured.iterdir() if p.is_file() and p.suffix.lower() in (".xlsx", ".xlsm", ".csv", ".tsv")),
                    key=lambda p: p.stat().st_mtime_ns,
                )
                if configured.is_dir()
                else ([configured] if configured.is_file() else [])
            )
            if files:
                for feed_file in files:
                    stat = feed_file.stat()
                    signature = (stat.st_mtime_ns, stat.st_size)
                    signature_key = str(feed_file.resolve())
                    if signature == last_signatures.get(signature_key):
                        continue
                    result = await asyncio.to_thread(
                        ingest_digital_file,
                        feed_file,
                        role=settings.digital_excel_role,
                        sheet_name=settings.digital_excel_sheet,
                    )
                    last_signatures[signature_key] = signature
                    logger.info("Sandbox 1.2 digital feed synchronized file={} result={}", feed_file.name, result)
            elif not configured.exists():
                logger.warning("Digital Excel feed not found: {}", configured)
        except asyncio.CancelledError:
            raise
        except Exception:
            logger.exception("Digital Excel feed synchronization failed")
        await asyncio.sleep(max(3.0, settings.digital_excel_poll_seconds))
