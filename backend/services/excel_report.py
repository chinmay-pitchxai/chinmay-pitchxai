import os
import re
import json
import sqlite3
import openpyxl
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths relative to this file
SERVICES_DIR = os.path.dirname(os.path.abspath(__file__))
BACKEND_DIR = os.path.dirname(SERVICES_DIR)
DATA_DIR = os.path.join(BACKEND_DIR, "data")
SOURCE_FILES_DIR = os.path.join(DATA_DIR, "source_files")

VERNIKA_DB = os.path.join(DATA_DIR, "vernika.db")
# Handle either naming scheme for doctors database
DOCTORS_DB = os.path.join(DATA_DIR, "testdataleads(doctors).db")
if not os.path.exists(DOCTORS_DB):
    DOCTORS_DB = os.path.join(DATA_DIR, "test_doctors.db")

def norm_phone(raw):
    """Normalize any phone number to 10-digit Indian mobile."""
    if raw is None: return None
    s = str(raw).strip()
    if s.endswith('.0'): s = s[:-2]
    s = re.sub(r'^\+', '', s)
    digits = re.sub(r'\D', '', s)
    if len(digits) > 10: digits = digits[-10:]
    return digits if len(digits) == 10 else None

def load_phones_from_excel(path, phone_col_idx):
    """Load unique 10-digit phones from a specific column in an Excel file."""
    if not os.path.exists(path):
        print(f"Excel source file not found: {path}")
        return set()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    phones = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue  # skip header
        if len(row) > phone_col_idx:
            p = norm_phone(row[phone_col_idx])
            if p: phones.add(p)
    wb.close()
    return phones

def get_lead_date(lead):
    from datetime import datetime, timezone, timedelta
    st = lead.get("start_time")
    if st and st > 0:
        try:
            tz = timezone(timedelta(hours=5, minutes=30))
            dt = datetime.fromtimestamp(st, tz=tz)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            pass
    return None

def get_report_data(from_date=None, to_date=None, role=None, include_transcripts: bool = True):
    """Load, classify, and deduplicate leads for KPI / Excel reports."""
    all_raw: list[dict] = []

    # Dashboard KPI for a campaign role: use only the live campaign DB (respects wipe/delete).
    if role:
        from core.storage import _get_conn, _row_to_dict
        from core.state import normalize_console_role

        rid = normalize_console_role(role)
        conn = _get_conn()
        try:
            rows = conn.execute("SELECT * FROM leads WHERE role = ?", (rid,)).fetchall()
            all_raw = [_row_to_dict(r) for r in rows]
        except Exception:
            all_raw = []
    else:
        # Legacy Excel export: merge historical source files + optional doctors DB.
        # Define exact source file paths
        LUX_CAR_PATH = os.path.join(SOURCE_FILES_DIR, "luxury car numbers .xlsx")
        CAR_DATA_PATH = os.path.join(SOURCE_FILES_DIR, "car data .xlsx")
        MYGATE_29_PATH = os.path.join(SOURCE_FILES_DIR, "mygate leads 29-jun.xlsx")
        MYGATE_03_PATH = os.path.join(SOURCE_FILES_DIR, "mygate leads 3-jul-2026.xlsx")

        # Load phone lists
        luxury_car_phones = load_phones_from_excel(LUX_CAR_PATH, phone_col_idx=1)
        car_phones        = load_phones_from_excel(CAR_DATA_PATH, phone_col_idx=1)
        mygate_29_phones  = load_phones_from_excel(MYGATE_29_PATH, phone_col_idx=1)
        mygate_03_phones  = load_phones_from_excel(MYGATE_03_PATH, phone_col_idx=3)

        all_mygate_phones = mygate_29_phones | mygate_03_phones
        recco_phones = {"9845471680", "9972229964", "9990362465", "9611344722"}

        def load_leads(db_path):
            # Main database is now PostgreSQL.
            if os.path.abspath(db_path) == os.path.abspath(VERNIKA_DB):
                try:
                    from core.storage import _get_conn

                    rows = _get_conn().execute("SELECT * FROM leads").fetchall()
                    return [dict(r) for r in rows]
                except Exception:
                    return []
            if not os.path.exists(db_path):
                return []
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            try:
                cur.execute("SELECT * FROM leads")
                rows = [dict(r) for r in cur.fetchall()]
            except Exception:
                rows = []
            conn.close()
            return rows

        v_leads = load_leads(VERNIKA_DB)
        d_leads = load_leads(DOCTORS_DB)

        v_phones = {norm_phone(l['phone']) for l in v_leads if norm_phone(l['phone'])}
        unique_docs = [l for l in d_leads if norm_phone(l['phone']) not in v_phones]

        all_raw = v_leads + unique_docs

    # Phone sets used for legacy classification (only when not role-scoped dashboard).
    if role:
        luxury_car_phones = set()
        car_phones = set()
        all_mygate_phones = set()
        recco_phones = set()
    else:
        LUX_CAR_PATH = os.path.join(SOURCE_FILES_DIR, "luxury car numbers .xlsx")
        CAR_DATA_PATH = os.path.join(SOURCE_FILES_DIR, "car data .xlsx")
        MYGATE_29_PATH = os.path.join(SOURCE_FILES_DIR, "mygate leads 29-jun.xlsx")
        MYGATE_03_PATH = os.path.join(SOURCE_FILES_DIR, "mygate leads 3-jul-2026.xlsx")
        luxury_car_phones = load_phones_from_excel(LUX_CAR_PATH, phone_col_idx=1)
        car_phones        = load_phones_from_excel(CAR_DATA_PATH, phone_col_idx=1)
        mygate_29_phones  = load_phones_from_excel(MYGATE_29_PATH, phone_col_idx=1)
        mygate_03_phones  = load_phones_from_excel(MYGATE_03_PATH, phone_col_idx=3)
        all_mygate_phones = mygate_29_phones | mygate_03_phones
        recco_phones = {"9845471680", "9972229964", "9990362465", "9611344722"}

    EXCLUDED_PHONES = {"9902578936", "7204955388"}
    EXCLUDED_NAMES  = {"chinmay"}

    # Filter by date if specified
    if from_date or to_date:
        filtered_raw = []
        for lead in all_raw:
            ldate = get_lead_date(lead)
            if not ldate:
                # If date filtering is active, skip leads with no resolved date (uncalled/pending)
                continue
            if from_date and ldate < from_date:
                continue
            if to_date and ldate > to_date:
                continue
            filtered_raw.append(lead)
        all_raw = filtered_raw

    def clean_source_name(name):
        if not name: return None
        name = re.sub(r'\.xlsx$|\.csv$', '', name, flags=re.IGNORECASE)
        name = re.sub(r'[\s_.-]+', ' ', name).strip()
        name_low = name.lower()
        if "recco" in name_low: return "Recco Data"
        if "mygate" in name_low or "my gate" in name_low: return "MyGate Data"
        if "luxury car" in name_low or "luxury_car" in name_low:
            if "new" in name_low: return "New Luxury Car Data"
            return "Luxury Car Data"
        return name.title() + " Data"

    def classify_lead(lead, phone):
        if phone in recco_phones: return "Recco Data"
        if phone in luxury_car_phones: return "Luxury Car Data"
        if phone in all_mygate_phones: return "MyGate Data"
        if phone in car_phones: return "New Luxury Car Data"
        
        try:
            ex = json.loads(lead.get("extra")) if isinstance(lead.get("extra"), (str, bytes, bytearray)) else (lead.get("extra") or {})
            if not isinstance(ex, dict):
                ex = {}
            source = ex.get("upload_source") or ex.get("Source File Name") or ex.get("source")
            if source:
                cleaned = clean_source_name(str(source))
                if cleaned: return cleaned
        except:
            pass
        return "Uncategorized"

    status_rank = {
        "site_visit": 7,
        "callback_completed": 6,
        "completed": 5,
        "callback_scheduled": 4,
        "not_interested": 3,
        "failed": 2,
        "pending": 1,
    }

    def get_db_key(l):
        status = (l.get("status") or "pending").lower()
        srank = status_rank.get(status, 1)
        try:
            ex = json.loads(l.get("extra") or "{}")
            has_outcome = 1 if (ex.get("Outcome") or ex.get("outcome") or
                                ex.get("disposition") or ex.get("call_outcome")) else 0
        except:
            has_outcome = 0
        updated = l.get("updated_at") or ""
        return (srank, has_outcome, updated)

    # Deduplicate by phone
    phone_best = {}
    for lead in all_raw:
        name  = (lead.get("name") or "").strip()
        phone = norm_phone(lead.get("phone"))
        if not phone: continue
        if phone in EXCLUDED_PHONES: continue
        if any(ex in name.lower() for ex in EXCLUDED_NAMES): continue

        k = get_db_key(lead)
        if phone not in phone_best or k > phone_best[phone][0]:
            phone_best[phone] = (k, lead)

    categories = {
        "Luxury Car Data":     [],
        "New Luxury Car Data": [],
        "MyGate Data":         [],
        "Recco Data":          [],
        "Doctors Data":        [],
    }

    for phone, (k, lead) in phone_best.items():
        cat = classify_lead(lead, phone)
        if cat not in categories:
            categories[cat] = []
        lead["_phone"] = phone
        categories[cat].append(lead)

    def get_extra(l):
        try: return json.loads(l.get("extra") or "{}")
        except: return {}

    def get_outcome(lead, extra):
        status = (lead.get("status") or "").lower()
        
        # 1. Try analysis JSON first (actual call outcome)
        try:
            an = json.loads(lead.get("analysis") or "{}")
            ao = (an.get("disposition") or an.get("outcome") or an.get("Outcome") or "").strip()
            if ao: return ao
        except: pass
        
        # 2. Fall back to uploaded extra fields
        outcome = (extra.get("Outcome") or extra.get("outcome") or
                   extra.get("disposition") or extra.get("call_outcome") or "").strip()
        if outcome: return outcome

        status_map = {
            "completed":           "Completed",
            "failed":              "Failed",
            "callback_scheduled":  "Callback Scheduled",
            "callback_completed":  "Callback Completed",
            "not_interested":      "Not Interested",
            "site_visit":          "Site Visit",
            "pending":             "Pending",
            "busy":                "Busy",
            "no-answer":           "No Answer",
            "no answer":           "No Answer",
            "canceled":            "Canceled",
        }
        return status_map.get(status, status.replace("_"," ").title())

    def build_row(lead, category):
        extra   = get_extra(lead)
        phone   = lead.get("_phone", "") or lead.get("phone", "") or ""
        name    = lead.get("name", "")
        status  = (lead.get("status") or "").lower()
        outcome = get_outcome(lead, extra) or ""
        out_low = outcome.lower()

        ldate = get_lead_date(lead)
        if ldate == "2026-07-14":
            if status == "site_visit" or "site visit" in out_low or "interested" in out_low:
                return None
            if "not interested" in out_low or "not_interested" in out_low or status == "not_interested":
                status = "no response"
                outcome = "No Response"
                out_low = "no response"

        # CRITICAL OVERRIDE: Nirmala.S (9448206173) is NOT a site visit
        if phone == "9448206173":
            status = "completed"
            outcome = "No Time"
            out_low = "no time"

        # CRITICAL OVERRIDE: RAJARSHI GUIN (8978090913) is NOT interested
        if phone == "8978090913":
            outcome = "Answered"
            out_low = "answered"

        call_made = 0 if status == "pending" else 1
        connected = 0
        failed_calls = 0
        interested = 0
        not_interested = 0
        no_response = 0
        voicemail = 0
        site_visit = 0
        callback = 0

        if call_made:
            if status in ("completed", "callback_scheduled", "callback_completed", "site_visit", "not_interested"):
                connected = 1
            else:
                connected = 0
                failed_calls = 1

            if connected:
                if status == "site_visit" or "site visit" in out_low or "site_visit" in out_low:
                    site_visit = 1
                elif "not interested" in out_low or "not_interested" in out_low or status == "not_interested":
                    not_interested = 1
                elif "voicemail" in out_low or "voice mail" in out_low:
                    voicemail = 1
                elif "no response" in out_low or "no_response" in out_low or "not reachable" in out_low or "no answer" in out_low or "busy" in out_low or out_low == "answered" or out_low == "completed" or out_low == "":
                    no_response = 1
                elif status in ("callback_scheduled", "callback_completed") or "call later" in out_low or "callback" in out_low:
                    callback = 1
                elif "interested" in out_low:
                    interested = 1
                else:
                    no_response = 1

        email = 1 if lead.get("email_sent") else 0
        wa = 1 if (site_visit == 1 or callback == 1 or interested == 1) else 0

        summary = ""
        rating = ""
        try:
            an = json.loads(lead.get("analysis") or "{}")
            summary = an.get("summary") or ""
            rating = an.get("rating") or ""
        except: pass

        if not summary:
            summary = extra.get("Summary") or extra.get("summary") or lead.get("details") or ""
        if not rating:
            rating = extra.get("Rating") or extra.get("rating") or ""

        # Format transcript from jsonl files (skip for live dashboard KPI — too slow on large lists)
        transcript = ""
        if include_transcripts:
            log_id = lead.get("_log_id") or ""
            role_val = lead.get("role") or ""
            if log_id and role_val:
                try:
                    from core.worker import _read_transcript_jsonl
                    raw_t = _read_transcript_jsonl(role_val, log_id)
                    if raw_t:
                        turns = []
                        for line in raw_t.splitlines():
                            if not line.strip(): continue
                            try:
                                obj = json.loads(line)
                                r = obj.get("role", "")
                                c = obj.get("content", "")
                                if r and c:
                                    turns.append(f"{r.capitalize()}: {c}")
                            except:
                                pass
                        transcript = "\n".join(turns)
                except Exception:
                    pass

        email_addr = lead.get("email") or ""
        if not email_addr:
            email_addr = extra.get("email_address") or extra.get("Email Id") or extra.get("Email") or ""

        return {
            "Category":       category,
            "Name":           name,
            "Phone":          lead.get("phone") or phone,
            "Role":           lead.get("role") or "",
            "Status":         status.replace("_"," ").title(),
            "Outcome":        outcome,
            "Call Made":      call_made,
            "Connected":      connected,
            "Failed Calls":   failed_calls,
            "Interested":     interested,
            "Not Interested": not_interested,
            "No Response":    no_response,
            "Voicemail":      voicemail,
            "Site Visit":     site_visit,
            "Callback":       callback,
            "WhatsApp Sent":  wa,
            "Email Sent":     email,
            "Email Address":  email_addr,
            "Rating":         rating,
            "Summary":        summary,
            "Transcript":     transcript,
            "Created At":     lead.get("created_at") or "",
            "Updated At":     lead.get("updated_at") or "",
        }

    STANDARD_ORDER = ["Luxury Car Data", "New Luxury Car Data", "MyGate Data", "Recco Data", "Doctors Data"]
    active_cats = [c for c in categories if len(categories[c]) > 0]
    
    CAT_ORDER = []
    for c in STANDARD_ORDER:
        if c in active_cats:
            CAT_ORDER.append(c)
    for c in sorted(active_cats):
        if c not in CAT_ORDER:
            CAT_ORDER.append(c)

    all_rows = []
    for cat in CAT_ORDER:
        for lead in categories[cat]:
            r_data = build_row(lead, cat)
            if r_data is not None:
                all_rows.append(r_data)

    return categories, all_rows, CAT_ORDER

def get_report_kpi_summary(from_date=None, to_date=None, role=None):
    """Get the KPI summary dictionary for dashboard display."""
    categories, all_rows, CAT_ORDER = get_report_data(
        from_date=from_date, to_date=to_date, role=role, include_transcripts=False
    )
    kpi = []
    for cat in CAT_ORDER:
        rows_cat = [r for r in all_rows if r["Category"] == cat]
        t  = len(rows_cat)
        cm = sum(r["Call Made"] for r in rows_cat)
        co = sum(r["Connected"] for r in rows_cat)
        fc = sum(r["Failed Calls"] for r in rows_cat)
        i  = sum(r["Interested"] for r in rows_cat)
        ni = sum(r["Not Interested"] for r in rows_cat)
        nr = sum(r["No Response"] for r in rows_cat)
        vm = sum(r["Voicemail"] for r in rows_cat)
        sv = sum(r["Site Visit"] for r in rows_cat)
        cb = sum(r["Callback"] for r in rows_cat)
        wa = sum(r["WhatsApp Sent"] for r in rows_cat)
        em = sum(r["Email Sent"] for r in rows_cat)
        kpi.append({
            "category": cat,
            "total_leads": t,
            "calls_made": cm,
            "connected": co,
            "failed_calls": fc,
            "interested": i,
            "not_interested": ni,
            "no_response": nr,
            "voicemail": vm,
            "site_visit": sv,
            "callback": cb,
            "whatsapp_sent": wa,
            "email_sent": em
        })
    
    # Calculate totals
    totals = {
        "category": "TOTAL",
        "total_leads": sum(x["total_leads"] for x in kpi),
        "calls_made": sum(x["calls_made"] for x in kpi),
        "connected": sum(x["connected"] for x in kpi),
        "failed_calls": sum(x["failed_calls"] for x in kpi),
        "interested": sum(x["interested"] for x in kpi),
        "not_interested": sum(x["not_interested"] for x in kpi),
        "no_response": sum(x["no_response"] for x in kpi),
        "voicemail": sum(x["voicemail"] for x in kpi),
        "site_visit": sum(x["site_visit"] for x in kpi),
        "callback": sum(x["callback"] for x in kpi),
        "whatsapp_sent": sum(x["whatsapp_sent"] for x in kpi),
        "email_sent": sum(x["email_sent"] for x in kpi)
    }

    return {"kpi": kpi, "totals": totals}

def generate_excel_report_workbook(from_date=None, to_date=None):
    """Build openpyxl workbook in memory and return its bytes."""
    categories, all_rows, CAT_ORDER = get_report_data(from_date=from_date, to_date=to_date)

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    CAT_COLORS = {
        "Luxury Car Data":     "7B2D00",
        "New Luxury Car Data": "1A237E",
        "MyGate Data":         "1B5E20",
        "Recco Data":          "4A148C",
        "Doctors Data":        "006064",
        "Inbound Calls":       "6366F1",
    }
    COLS = ["Category","Name","Phone","Role","Status","Outcome","Call Made","Connected",
            "Failed Calls","Interested","Not Interested","No Response","Voicemail","Site Visit",
            "Callback","WhatsApp Sent","Email Sent","Rating","Summary","Transcript","Created At","Updated At"]
    COL_W = [20,30,16,12,14,18,10,10,12,10,14,12,10,10,10,14,12,8,45,45,20,20]

    def mfill(h): return PatternFill(start_color=h, end_color=h, fill_type="solid")
    def mbord():
        t = Side(style="thin", color="DDDDDD")
        return Border(left=t, right=t, top=t, bottom=t)

    kpi = []

    for cat in CAT_ORDER:
        rows_cat = [r for r in all_rows if r["Category"] == cat]
        # Exclude Pending leads from sheet tabs
        rows_cat_called = [r for r in rows_cat if r["Status"] != "Pending"]
        
        ws = wb_out.create_sheet(title=cat[:31])
        hc = CAT_COLORS.get(cat, "263238")

        for ci, col in enumerate(COLS, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.fill = mfill(hc)
            c.font = Font(bold=True, color="FFFFFF", size=11)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = mbord()

        alt = mfill("F5F5F5")
        for ri, rd in enumerate(rows_cat_called, 2):
            for ci, col in enumerate(COLS, 1):
                val = rd.get(col,"")
                if col == "Email Sent" and rd.get("Email Sent") and rd.get("Email Address"):
                    val = f"Yes ({rd.get('Email Address')})"
                elif col == "WhatsApp Sent":
                    val = "Yes" if rd.get("WhatsApp Sent") == 1 else "No"
                elif col == "Email Sent" and not rd.get("Email Sent"):
                    val = "No"
                elif col == "Email Sent" and rd.get("Email Sent") and not rd.get("Email Address"):
                    val = "Yes"
                
                c = ws.cell(row=ri, column=ci, value=val)
                c.alignment = Alignment(vertical="center", wrap_text=(col=="Summary" or col=="Email Sent"))
                c.border = mbord()
                if ri % 2 == 0: c.fill = alt

        for i, w in enumerate(COL_W, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

        t  = len(rows_cat)
        cm = sum(r["Call Made"] for r in rows_cat)
        co = sum(r["Connected"] for r in rows_cat)
        fc = sum(r["Failed Calls"] for r in rows_cat)
        i  = sum(r["Interested"] for r in rows_cat)
        ni = sum(r["Not Interested"] for r in rows_cat)
        nr = sum(r["No Response"] for r in rows_cat)
        vm = sum(r["Voicemail"] for r in rows_cat)
        sv = sum(r["Site Visit"] for r in rows_cat)
        cb = sum(r["Callback"] for r in rows_cat)
        wa = sum(r["WhatsApp Sent"] for r in rows_cat)
        em = sum(r["Email Sent"] for r in rows_cat)
        kpi.append([cat, t, cm, co, fc, i, ni, nr, vm, sv, cb, wa, em])

    # Summary Sheet
    ws_s = wb_out.create_sheet(title="SUMMARY", index=0)
    sh = ["Category","Total Leads","Calls Made","Connected","Failed Calls","Interested",
          "Not Interested","No Response","Voicemail","Site Visit","Callback",
          "WhatsApp Sent","Email Sent"]
    sw = [25,13,12,12,12,12,14,12,10,10,10,14,12]

    for ci, h in enumerate(sh, 1):
        c = ws_s.cell(row=1, column=ci, value=h)
        c.fill = mfill("0D1B2A")
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = mbord()

    bg = list(CAT_COLORS.values())
    for ri, kr in enumerate(kpi, 2):
        for ci, val in enumerate(kr, 1):
            c = ws_s.cell(row=ri, column=ci, value=val)
            if ci == 1:
                c.fill = mfill(bg[(ri-2) % len(bg)])
                c.font = Font(bold=True, color="FFFFFF", size=11)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = mbord()

    totals = ["TOTAL"] + [sum(r[i] for r in kpi) for i in range(1, len(sh))]
    for ci, val in enumerate(totals, 1):
        c = ws_s.cell(row=len(kpi)+2, column=ci, value=val)
        c.fill = mfill("B71C1C")
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = mbord()

    for i, w in enumerate(sw, 1):
        ws_s.column_dimensions[get_column_letter(i)].width = w
    ws_s.row_dimensions[1].height = 35
    ws_s.freeze_panes = "A2"

    out_bio = BytesIO()
    wb_out.save(out_bio)
    out_bio.seek(0)
    return out_bio.getvalue()
